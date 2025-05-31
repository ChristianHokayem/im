import streamlit as st
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
import base64

# Load the base template Excel file
def load_template():
    return load_workbook("TEST Intelligences Multiples_Locked.xlsx")

# Extract questions from the questionnaire sheet
def get_questions():
    wb = load_template()
    sheet = wb["QUESTIONNAIRE"]
    questions = []
    for row in sheet.iter_rows(min_row=6, values_only=True):
        q1 = row[0]
        q2 = row[3]
        if q1:
            questions.append(q1.strip())
        if q2:
            questions.append(q2.strip())
    return questions

# Write user answers to the Excel copy
def fill_excel(name, family_name, answers):
    wb = load_template()
    sheet = wb["QUESTIONNAIRE"]
    answer_index = 0
    for row in sheet.iter_rows(min_row=6, max_row=6+len(answers)//2):
        if answer_index >= len(answers):
            break
        if answers[answer_index]:
            row[1].value = 1
        answer_index += 1
        if answer_index < len(answers) and answers[answer_index]:
            row[4].value = 1
        answer_index += 1

    # Save to BytesIO
    output = BytesIO()
    filename = f"{family_name}_{name}_Intelligences.xlsx"
    wb.save(output)
    output.seek(0)
    return output, filename

# Display download link
def get_download_link(buffer, filename):
    b64 = base64.b64encode(buffer.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download your result file</a>'
    return href

# Streamlit UI
st.title("Test des intelligences multiples")
st.write("Veuillez cocher les affirmations qui vous correspondent.")

name = st.text_input("Prénom")
family_name = st.text_input("Nom de famille")

questions = get_questions()
answers = []

st.subheader("Questionnaire")
for q in questions:
    answers.append(st.checkbox(q))

if st.button("Soumettre"):
    if not name or not family_name:
        st.warning("Veuillez entrer votre prénom et votre nom de famille.")
    else:
        file_buffer, filename = fill_excel(name, family_name, answers)
        st.success("Formulaire rempli avec succès !")
        st.markdown(get_download_link(file_buffer, filename), unsafe_allow_html=True)
